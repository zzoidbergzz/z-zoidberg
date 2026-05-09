#!/usr/bin/env python3
"""Import MITRE ATT&CK threat actor encyclopedia from Google Sheets.

Maps every APT group across vendor taxonomies (CrowdStrike, Kaspersky,
Mandiant, etc.), their operations, tooling, targets, and source links.
"""
import json
import os
import sys
import time
import urllib.request

import openpyxl

API = "http://localhost:8010"

def get_api_key():
    # First check env
    key = os.environ.get("SK_API_KEY", "")
    if key:
        return key
    # Try reading from bootstrap password
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("BOOTSTRAP_ADMIN_PASSWORD="):
                    pwd = line.split("=", 1)[1].strip().strip('"')
                    # Need to get a JWT token using this password
                    try:
                        import urllib.parse
                        data = urllib.parse.urlencode({"username": "m@z.je", "password": pwd}).encode()
                        req = urllib.request.Request(f"{API}/api/v1/auth/token", data=data)
                        with urllib.request.urlopen(req, timeout=10) as resp:
                            d = json.loads(resp.read())
                            return d.get("access_token", pwd)
                    except:
                        return pwd
    return ""

KEY = os.environ.get("SK_API_KEY", get_api_key())

def mcp_call(tool, args, timeout=30):
    data = json.dumps({"tool": tool, "args": args}).encode()
    req = urllib.request.Request(f"{API}/api/v1/mcp/call", data=data, headers={
        "X-API-Key": KEY, "Content-Type": "application/json"
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode()[:500]
        return {"error": f"HTTP {e.code}: {body}"}
    except Exception as e:
        return {"error": str(e)}


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def import_sheet(wb, sheet_name, country):
    """Import a country sheet of threat actors."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return 0

    headers = [safe_str(h) for h in rows[1]]  # Row 2 is headers
    imported = 0

    for row in rows[2:]:  # Data starts at row 3
        values = [safe_str(v) for v in row]
        if not values[0] or len(values[0]) < 2:
            continue

        common_name = values[0]
        # Skip header-like rows
        if common_name.lower() in ('common name', 'name', 'comment', 'total', ''):
            continue

        # Find key columns (varies per sheet)
        aliases = []
        mitre_id = ""
        operations = []
        tooling = ""
        targets = ""
        modus = ""
        comment = ""
        links = []

        for i, h in enumerate(headers):
            if i >= len(values):
                break
            v = values[i]
            if not v:
                continue
            h_lower = h.lower()

            # Aliases
            if any(x in h_lower for x in ['other name', 'crowdstrike', 'kaspersky',
                                           'secureworks', 'mandiant', 'fireeye',
                                           'symantec', 'isight', 'cisco', 'palo alto',
                                           'talos', 'irl', 'rep. of korea', 'naming',
                                           'nsa', 'fbi']):
                if v and v != common_name and len(v) < 100:
                    aliases.append(v)

            # MITRE
            if 'mitre' in h_lower and 'att&ck' in h_lower:
                if v and v.startswith('G'):
                    mitre_id = v.split(',')[0].strip()

            # Operations
            if 'operation' in h_lower:
                if v and len(v) < 200:
                    operations.append(v)

            # Tooling
            if 'toolset' in h_lower or 'malware' in h_lower:
                if v:
                    tooling = v[:2000]

            # Targets
            if 'target' in h_lower:
                targets = v[:2000]

            # Modus operandi
            if 'modus' in h_lower:
                modus = v[:2000]

            # Comment
            if 'comment' in h_lower:
                comment = v[:2000]

            # Links
            if 'link' in h_lower and v.startswith('http'):
                links.append(v)

        # Create entity
        description = f"Threat actor attributed to {country}."
        if comment:
            description = comment[:500]
        if targets:
            description += f" Targets: {targets[:200]}"

        result = mcp_call("create_entity", {
            "name": common_name,
            "kind": "threat_actor",
        })

        entity_id = result.get("id")
        if not entity_id and "error" in result:
            # Entity likely exists, search for it
            try:
                req2 = urllib.request.Request(
                    f"{API}/api/v1/entities/?limit=500",
                    headers={"X-API-Key": KEY}
                )
                with urllib.request.urlopen(req2, timeout=30) as resp2:
                    ents = json.loads(resp2.read())
                    for e in ents:
                        if e.get("canonical_name") == common_name:
                            entity_id = e["id"]
                            break
            except:
                pass

        if not entity_id:
            print(f"  SKIP: {common_name} (no entity_id)")
            continue

        # Add attribution claim
        aliases_clean = [a for a in aliases if a and len(a) > 1]
        mcp_call("create_claim", {
            "entity_id": entity_id,
            "claim_type": "attribution",
            "value": {
                "text": f"Threat actor attributed to {country}. Known aliases: {', '.join(aliases_clean[:10])}.",
                "country": country,
                "aliases": aliases_clean[:15],
                **({"mitre_attack_id": mitre_id} if mitre_id else {}),
            },
            "confidence": 0.8,
        })

        # Add operations
        ops_clean = [o for o in operations if o and len(o) > 1]
        if ops_clean:
                mcp_call("create_claim", {
                    "entity_id": entity_id,
                    "claim_type": "campaign",
                    "value": {
                        "text": f"Known operations: {', '.join(ops_clean[:8])}.",
                        "operations": ops_clean[:10],
                    },
                    "confidence": 0.85,
                })

        # Add tooling
        if tooling:
            mcp_call("create_claim", {
                "entity_id": entity_id,
                "claim_type": "tooling",
                "value": {
                    "text": f"Known toolset/malware: {tooling[:500]}",
                    "tools": tooling[:1000],
                },
                "confidence": 0.8,
            })

        # Add targets
        if targets:
            mcp_call("create_claim", {
                "entity_id": entity_id,
                "claim_type": "infrastructure",
                "value": {
                    "text": f"Known targets: {targets[:500]}",
                    "targets": targets[:1000],
                },
                "confidence": 0.8,
            })

        # Add TTPs
        if modus:
            mcp_call("create_claim", {
                "entity_id": entity_id,
                "claim_type": "technique",
                "value": {
                    "text": f"Modus operandi: {modus[:500]}",
                    "ttps": modus[:1000],
                },
                "confidence": 0.7,
            })

        # Add source links
        if links:
            mcp_call("create_claim", {
                "entity_id": entity_id,
                "claim_type": "evidence",
                "value": {
                    "text": f"Sources: {len(links)} references",
                    "links": links[:25],
                },
                "confidence": 1.0,
            })

        imported += 1
        print(f"  + {common_name} ({country}) — {len(aliases_clean)} aliases, {len(ops_clean)} ops, {len(links)} sources")

    return imported


def main():
    xlsx = "/tmp/mitre-attack-data.xlsx"
    if not os.path.exists(xlsx):
        print(f"ERROR: {xlsx} not found")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True)

    country_sheets = {
        "China": "China",
        "Russia": "Russia",
        "North Korea": "North Korea",
        "Iran": "Iran",
        "Israel": "Israel",
        "NATO": "NATO",
        "Middle East": "Middle East",
        "Others": "Other",
        "Unknown": "Unknown",
    }

    total = 0
    for sheet_name, country in country_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        print(f"\n=== {country} ===")
        count = import_sheet(wb, sheet_name, country)
        total += count
        print(f"  Imported {count} actors from {sheet_name}")

    print(f"\nTotal: {total} threat actors imported")


if __name__ == "__main__":
    main()
